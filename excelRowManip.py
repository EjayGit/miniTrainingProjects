from openpyxl import Workbook

newWookbook = Workbook()
sheet = newWookbook.active
sheet["A1"] = "name"
sheet["B1"] = "surname"
sheet.insert_cols(idx=1)
sheet.insert_cols(idx=3, amount=5)
sheet.delete_cols(idx=3)

newWookbook.save("excelRowManipOut.xlsx")

for row in sheet.iter_rows(values_only=True):
    print(row)
sheet.insert_rows(idx=1, amount=5)
sheet.delete_rows(idx=3, amount=2)
for row in sheet.iter_rows(values_only=True):
    print(row)