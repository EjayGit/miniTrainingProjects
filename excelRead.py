from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pprint

# Automate Daily Routine Excel Task

"""The program will do the followings.
- Read the data from Excel spreadsheet
- Counts the number of transactions for each supplier
- Sum the transaction amount for each supplier
- Prints the result to the custom file
## Steps for development"""

# - TODO 1 - Open and read the cells of an Excel document with openpyxl module
workbook = load_workbook("transactions.xlsx")
sheet = workbook.active
supplierData= {}
for row in range(2,sheet.max_row+1):
    # For each column, set data for each col in row
    transactionType = sheet['B'+str(row)].value
    supplierName = sheet['C'+str(row)].value
    amount = sheet['D'+str(row)].value
    #print(transactionType)
    #print(supplierName)
    #print(amount)
    # - TODO 2 - Calculate all the transaction amounts and store it in dictionary data structure
    supplierData.setdefault(transactionType,{})
    supplierData[transactionType].setdefault(supplierName,{'transactionCount':0,
                                                            'amount':0})
    supplierData[transactionType][supplierName]['transactionCount'] += 1
    supplierData[transactionType][supplierName]['amount'] += int(amount)


# - TODO 3 - Write the data structure to .py file using pprint module
with open('projectOut.py', 'w') as file:
    file.write(pprint.pformat(supplierData))
