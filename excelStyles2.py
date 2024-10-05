from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule, IconSetRule
from openpyxl import load_workbook


spreadsheet = load_workbook("name.xlsx")
sheet = spreadsheet.active
colorScalerule = ColorScaleRule(start_type="min",
                                start_color="00FF0000",
                                mid_type="num",
                                mid_value=60,
                                mid_color="00FFFF00",
                                end_type="max",
                                end_color="0000FF00")
sheet.conditional_formatting.add("B2:B100", colorScalerule)
spreadsheet.save("name3.xlsx")