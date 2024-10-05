from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule, IconSetRule, DataBarRule
from openpyxl import load_workbook


spreadsheet = load_workbook("name.xlsx")
sheet = spreadsheet.active
iconSetRule = IconSetRule("3Arrows", "num",[40, 60, 81])
dataBarRule = DataBarRule(start_type="num",
                            start_value=30,
                            end_type="num",
                            end_value=70,
                            color="0000FF00")
sheet.conditional_formatting.add("C2:C100", iconSetRule)
sheet.conditional_formatting.add("B2:B100", dataBarRule)
spreadsheet.save("name4.xlsx")