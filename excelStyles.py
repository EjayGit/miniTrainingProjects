from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule
from openpyxl import load_workbook


spreadsheet = load_workbook("name.xlsx")
sheet = spreadsheet.active
red_background = PatternFill(bgColor="00FF0000")
differential_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=differential_style)
rule.formula = ["$B1<50"]
sheet.conditional_formatting.add("A1:B100", rule)
spreadsheet.save("name2.xlsx")