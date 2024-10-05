from openpyxl import Workbook

# open new and save workbook
newWorkbook = Workbook()
sheet = newWorkbook.active
sheet["A1"] = "name"
sheet["A2"] = "Erik"
sheet["B1"] = "surname"
sheet["B2"] = "Clark"
sheet["B11"] = "test"

cell = sheet["A2"]
print(cell.value)
cell.value = "Tom"
for row in sheet.iter_rows(values_only=True):
    print(row)
newWorkbook.save('excelWriteProj.xlsx')

newWorkbook.create_sheet("New Sheet")
print(newWorkbook.sheetnames)
print(sheet.title)
newSheet = newWorkbook["New Sheet"]
newSheet["A1"] = "ID"
newSheet["B1"] = 'Address'
newSheet["C14"] = "new sheet testing"
newSheet["D1"] = 'newtesting'
sheet["C1"] = "sheetTest"
newSheet["E1"] = 'newTest'
newWorkbook.remove(sheet)
newWorkbook.copy_worksheet(newSheet)
print(newWorkbook.sheetnames)
newWorkbook.save("newSpreadsheet.xlsx")