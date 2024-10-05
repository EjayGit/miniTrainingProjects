from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# TODO 1 - Open current excel file and create new file for output. Create two list for each column
workbookIn = load_workbook("sample_data.xlsx")
workbookOut = Workbook()
sheetIn = workbookIn.active
sheetOut = workbookOut.active
rows=sheetIn.max_row
cols=sheetIn.max_column

# TODO 2 - Create nested loop to loop through rows and columns. Insert column 1 values to col_1 and column 2 values to col_2
for row in range(1,rows+1):
    for col in range(1,cols+1):
        sheetOut[get_column_letter(row)+str(col)] = sheetIn[get_column_letter(col)+str(row)].value

# TODO 3 - To transpose we need to loop through each list and insert values to rows of new excel file and save output
workbookOut.save("excelTransposeOut.xlsx")