from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import datetime as dt
import random


# TODO: Generate custome data and save to excel.abs
spreadsheet = Workbook()
sheet = spreadsheet.active
sheet["A2"] = "Paperback"
sheet["A3"] = "Ebook"
for month in range(2,14):
    sheet[get_column_letter(month)+"1"] = dt.datetime(2025,month-1,13).strftime("%B")
    sheet[get_column_letter(month)+"2"] = random.randint(15, 40)
    sheet[get_column_letter(month)+"3"] = random.randint(15, 40)

# TODO: Create a line graph
chart=LineChart()
data = Reference(worksheet=sheet, max_col=13, min_col=1, min_row=2, max_row=3)
chart.add_data(data, from_rows=True, titles_from_data=True)
sheet.add_chart(chart,"B5")

# TODO: Create generator for the graph
cats = Reference(worksheet=sheet, min_col=2, max_col=13, min_row=1, max_row=1)
chart.set_categories(cats)
spreadsheet.save('chartProj.xlsx')