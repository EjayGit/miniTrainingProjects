from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pprint

workbook = load_workbook("AmazonReviews.xlsx")
sheet = workbook.active
"""print(sheet.title)
cell1 = sheet["B2"]
cell2 = sheet.cell(row=2, column=3) # cell numbers start at 1, not 0.
print(cell1.value)
print(cell2.value)
#colRange = sheet["A"] # ["A":"C"]
#print(colRange)
rowRange = sheet["1"]
print(rowRange)
for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10, values_only=True):
    print(row)
for col in sheet.iter_cols(min_row=1, max_row=10, min_col=1, max_col=10, values_only=True):
    print(col)
for cols in sheet.columns:
    for col in cols:
        print(col.value)"""

for cells in sheet["A1":"A3"]:
    for cell in cells:
        print(cell.coordinate, cell.value)

"""print(get_column_letter(sheet.max_column))
print(column_index_from_string("J"))
for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    print(value)

reviews={}
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=8, values_only=True):
    print(row)
    reviewID=row[0]
    review = {
        "profileName": row[1],
        "title":row[4],
        "rating":row[5]
    }
    reviews[reviewID] = review
pprint.pprint(reviews)"""